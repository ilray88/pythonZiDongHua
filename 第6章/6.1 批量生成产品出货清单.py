from openpyxl import load_workbook
workbook = load_workbook('F:\\代码文件\\第6章\\出货统计表.xlsx')
worksheet = workbook['Sheet1']
data = {}
for row in range(2, worksheet.max_row + 1):
    date = worksheet['B' + str(row)].value.date()
    customer = worksheet['C' + str(row)].value
    product = worksheet['D' + str(row)].value
    number = worksheet['E' + str(row)].value
    model = worksheet['G' + str(row)].value
    info_list = [customer, product, number, model]
    data.setdefault(date, [])
    data[date].append(info_list)
for key, value in data.items():
    print(key, value)
workbook_day = load_workbook('F:\\代码文件\\第6章\\出货清单模板.xlsx')
worksheet_day = workbook_day['出货清单模板']
for date in data.keys():
    worksheet_new = workbook_day.copy_worksheet(worksheet_day)
    worksheet_new.title = str(date)[-5:]
    worksheet_new.cell(row=2, column=5).value = date
    i = 4
    for product in data[date]:
        worksheet_new.cell(row=i, column=2).value = product[0]
        worksheet_new.cell(row=i, column=3).value = product[1]
        worksheet_new.cell(row=i, column=4).value = product[2]
        worksheet_new.cell(row=i, column=5).value = product[3]
        i += 1
workbook_day.save('F:\\代码文件\\第6章\\产品出货清单.xlsx')
